"""
compare_all.py  —  V4 Edition
Trains and compares all four VMamba variants and saves input/reconstructed images.

Usage
─────
  cd version4
  python comparison/compare_all.py --epochs 5 --batch_size 1 --patch_size 32 --d_model 32
  python comparison/compare_all.py --epochs 30 --batch_size 1 --patch_size 32 --d_model 32
  python comparison/compare_all.py --skip_old --skip_improved --skip_v3   # V4 only
"""

import sys
import os
import json
import time
import math
import argparse
import datetime

_HERE    = os.path.dirname(os.path.abspath(__file__))
_V4      = os.path.dirname(_HERE)
_PROJECT = os.path.dirname(_V4)
sys.path.insert(0, os.path.join(_PROJECT, 'version3'))  # lowest priority
sys.path.insert(0, _PROJECT)
sys.path.insert(0, os.path.join(_PROJECT, 'scripts'))
sys.path.insert(0, _V4)                                  # highest priority — version4 baseline_models wins

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
import numpy as np

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    HAS_XL = True
except ImportError:
    HAS_XL = False

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(x, **kw): return x

from dataset_loader_overlap import create_dataloaders_overlap
from baseline_models import create_vmamba_model, count_parameters
from loss_functions_v4 import CompositeLossV4, compute_psnr, compute_sam_metric, compute_ergas
from loss_functions import compute_ssim

# ── Paths ─────────────────────────────────────────────────────────────────────
CHIKUSEI_PATH = os.path.join(_PROJECT, 'chikusei', 'chikusei.mat')
CKPT_DIR      = os.path.join(_HERE, 'checkpoints')
PLOTS_DIR     = os.path.join(_HERE, 'plots')
RESULTS_DIR   = os.path.join(_HERE, 'results')
IMAGES_DIR    = os.path.join(_HERE, 'saved_images')

for d in [CKPT_DIR, PLOTS_DIR, RESULTS_DIR, IMAGES_DIR]:
    os.makedirs(d, exist_ok=True)

IN_CHANNELS = 128
SCALE       = 4


# ============================================================================
# Excel Logger
# ============================================================================

class ExcelLogger:
    """Logs per-epoch metrics and final summary to .xlsx."""

    EPOCH_COLS = [
        ('Model', 14), ('Epoch', 7), ('Train Loss', 12),
        ('Val PSNR (dB)', 14), ('Val SAM (deg)', 14),
        ('Val ERGAS', 12), ('Learning Rate', 14),
        ('Epoch Time (s)', 14), ('Best?', 7),
    ]

    MODEL_FILLS = {
        'old':      PatternFill('solid', fgColor='BDD7EE') if HAS_XL else None,  # blue
        'improved': PatternFill('solid', fgColor='FFEB9C') if HAS_XL else None,  # yellow
        'v3':       PatternFill('solid', fgColor='C6EFCE') if HAS_XL else None,  # green
        'v4':       PatternFill('solid', fgColor='E2CFFF') if HAS_XL else None,  # purple
    }

    def __init__(self, path: str):
        self.path = path
        if not HAS_XL:
            return
        self.wb = openpyxl.Workbook()

        # Epoch Log sheet
        self.ws_log = self.wb.active
        self.ws_log.title = 'Epoch Log'
        hdrs = [c[0] for c in self.EPOCH_COLS]
        widths = [c[1] for c in self.EPOCH_COLS]
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            cell = self.ws_log.cell(1, ci, h)
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill      = PatternFill('solid', fgColor='4F81BD')
            self.ws_log.column_dimensions[get_column_letter(ci)].width = w

        # Summary sheet
        self.ws_sum = self.wb.create_sheet('Summary')
        self._init_summary_sheet()

    def _init_summary_sheet(self):
        hdrs = ['Metric', 'Old VMamba', 'Improved VMamba', 'V3 VMamba', 'V4 VMamba']
        for ci, h in enumerate(hdrs, 1):
            cell = self.ws_sum.cell(1, ci, h)
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill      = PatternFill('solid', fgColor='4F81BD')
        widths = [20, 16, 16, 14, 14]
        for ci, w in enumerate(widths, 1):
            self.ws_sum.column_dimensions[get_column_letter(ci)].width = w

    def log_epoch(self, model_name, epoch, train_loss,
                  val_psnr, val_sam, val_ergas, lr, t, is_best):
        if not HAS_XL:
            return
        row = [model_name, epoch, round(train_loss, 6),
               round(val_psnr, 4), round(val_sam, 4),
               round(val_ergas, 4), f'{lr:.2e}',
               round(t, 1), '*' if is_best else '']
        ri = self.ws_log.max_row + 1
        fill = self.MODEL_FILLS.get(model_name)
        for ci, v in enumerate(row, 1):
            cell = self.ws_log.cell(ri, ci, v)
            cell.alignment = Alignment(horizontal='center')
            if fill:
                cell.fill = fill
            if is_best and ci == len(row):
                cell.font = Font(bold=True, color='FF0000')
        self.wb.save(self.path)

    def write_summary(self, best: dict):
        if not HAS_XL:
            return
        metrics = [
            ('Best PSNR (dB)',  'psnr',  True),
            ('Best SAM (deg)',   'sam',   False),
            ('Best ERGAS',       'ergas', False),
            ('Params (M)',       'params', False),
        ]
        variants = list(best.keys())
        GREEN = PatternFill('solid', fgColor='C6EFCE')
        RED   = PatternFill('solid', fgColor='FFC7CE')
        for ri, (label, key, higher_better) in enumerate(metrics, 2):
            self.ws_sum.cell(ri, 1, label).font = Font(bold=True)
            vals = [best[v].get(key, float('nan')) for v in variants]
            try:
                best_val = max(vals) if higher_better else min(vals)
            except Exception:
                best_val = None
            for ci, (v, val) in enumerate(zip(variants, vals), 2):
                cell = self.ws_sum.cell(ri, ci, round(val, 4) if not math.isnan(val) else 'N/A')
                cell.alignment = Alignment(horizontal='center')
                if best_val is not None and not math.isnan(val):
                    cell.fill = GREEN if val == best_val else RED
        self.wb.save(self.path)


# ============================================================================
# Image Saver
# ============================================================================

def save_images(batch, outputs: dict, epoch: int, tag: str = ''):
    """
    Save LR-HSI (false-colour), HR-PAN, GT HR-HSI, and each model's output.
    Uses bands [30, 20, 10] as RGB proxy for false-colour composite.

    Args:
        batch:   dict with 'lr_hsi', 'hr_pan', 'hr_hsi' tensors (B=1)
        outputs: {variant: tensor (1,128,H,W)} predicted HR-HSI
        epoch:   current epoch number
        tag:     filename suffix
    """
    if not HAS_MPL:
        return

    def to_rgb(hsi_tensor):
        """Convert HSI (1,C,H,W) → RGB numpy using bands [30,20,10]."""
        t = hsi_tensor[0]       # (C, H, W)
        r = t[min(30, t.shape[0]-1)].cpu().numpy()
        g = t[min(20, t.shape[0]-1)].cpu().numpy()
        b = t[min(10, t.shape[0]-1)].cpu().numpy()
        rgb = np.stack([r, g, b], axis=2)
        rgb = np.clip(rgb, 0, 1)
        return rgb

    def to_gray(pan_tensor):
        return pan_tensor[0, 0].cpu().numpy()

    lr_hsi = batch['lr_hsi']
    hr_pan = batch['hr_pan']
    hr_hsi = batch['hr_hsi']

    variant_names = list(outputs.keys())
    n_cols = 3 + len(variant_names)   # LR-HSI, HR-PAN, GT, + each model
    fig, axes = plt.subplots(1, n_cols, figsize=(4 * n_cols, 4))

    # LR-HSI upsampled for display
    lr_up = torch.nn.functional.interpolate(
        lr_hsi, scale_factor=SCALE, mode='bicubic', align_corners=False)
    axes[0].imshow(to_rgb(lr_up), aspect='auto')
    axes[0].set_title('LR-HSI\n(Bicubic up)', fontsize=9)
    axes[0].axis('off')

    axes[1].imshow(to_gray(hr_pan), cmap='gray', aspect='auto')
    axes[1].set_title('HR-PAN', fontsize=9)
    axes[1].axis('off')

    axes[2].imshow(to_rgb(hr_hsi), aspect='auto')
    axes[2].set_title('GT HR-HSI', fontsize=9)
    axes[2].axis('off')

    for ci, (vname, pred) in enumerate(outputs.items(), 3):
        err = (pred - hr_hsi).abs().mean(dim=1, keepdim=True)   # (1,1,H,W)
        err_np = err[0, 0].cpu().numpy()
        # Show false-colour of prediction
        axes[ci].imshow(to_rgb(pred.clamp(0, 1)), aspect='auto')
        axes[ci].set_title(f'{vname.upper()}\nPSNR={compute_psnr(pred, hr_hsi):.2f}dB',
                           fontsize=8)
        axes[ci].axis('off')

    plt.suptitle(f'Epoch {epoch} — False Colour RGB (bands 30,20,10)', fontsize=10)
    plt.tight_layout()
    fname = os.path.join(IMAGES_DIR, f'epoch{epoch:04d}_{tag}.png')
    plt.savefig(fname, dpi=120, bbox_inches='tight')
    plt.close()

    # Also save individual error maps
    fig2, axes2 = plt.subplots(1, len(variant_names), figsize=(4*len(variant_names), 4))
    if len(variant_names) == 1:
        axes2 = [axes2]
    for ax, (vname, pred) in zip(axes2, outputs.items()):
        err = (pred - hr_hsi).abs().mean(dim=1)[0].cpu().numpy()
        im = ax.imshow(err, cmap='hot', aspect='auto')
        ax.set_title(f'{vname.upper()} Error\n(mean abs per pixel)', fontsize=8)
        ax.axis('off')
        plt.colorbar(im, ax=ax, fraction=0.046)
    plt.suptitle(f'Epoch {epoch} — Spectral-Mean Absolute Error', fontsize=10)
    plt.tight_layout()
    fname2 = os.path.join(IMAGES_DIR, f'epoch{epoch:04d}_{tag}_error.png')
    plt.savefig(fname2, dpi=120, bbox_inches='tight')
    plt.close()


def save_spectral_curves(batch, outputs: dict, epoch: int, pixel=(16, 16)):
    """Save spectral profile at a given pixel for GT vs all models."""
    if not HAS_MPL:
        return
    hr_hsi = batch['hr_hsi']
    py, px = pixel
    H, W = hr_hsi.shape[2], hr_hsi.shape[3]
    py, px = min(py, H-1), min(px, W-1)

    gt_spec = hr_hsi[0, :, py, px].cpu().numpy()

    plt.figure(figsize=(10, 4))
    plt.plot(gt_spec, 'k-', linewidth=2, label='GT')
    colours = ['blue', 'orange', 'green', 'purple']
    for (vname, pred), colour in zip(outputs.items(), colours):
        pred_spec = pred[0, :, py, px].cpu().detach().numpy()
        plt.plot(pred_spec, '--', color=colour, label=vname.upper())
    plt.xlabel('Band index')
    plt.ylabel('Reflectance')
    plt.title(f'Epoch {epoch} — Spectral curve at pixel ({py},{px})')
    plt.legend()
    plt.tight_layout()
    fname = os.path.join(IMAGES_DIR, f'epoch{epoch:04d}_spectral_curve.png')
    plt.savefig(fname, dpi=100, bbox_inches='tight')
    plt.close()


# ============================================================================
# Training
# ============================================================================

def train_one_epoch(model, loader, criterion, optimiser, device):
    model.train()
    total_loss = 0.0
    for batch in tqdm(loader, desc='    train', leave=False):
        lr_hsi = batch['lr_hsi'].to(device)
        hr_pan = batch['hr_pan'].to(device)
        hr_hsi = batch['hr_hsi'].to(device)
        optimiser.zero_grad()
        pred = model(lr_hsi, hr_pan)
        loss = criterion(pred, hr_hsi)
        loss.backward()
        nn.utils.clip_grad_norm_(model.parameters(), 1.0)
        optimiser.step()
        total_loss += loss.item()
    return total_loss / max(len(loader), 1)


@torch.no_grad()
def validate(model, loader, device):
    model.eval()
    ps, ss, es, si = [], [], [], []
    for batch in loader:
        lr_hsi = batch['lr_hsi'].to(device)
        hr_pan = batch['hr_pan'].to(device)
        hr_hsi = batch['hr_hsi'].to(device)
        pred   = model(lr_hsi, hr_pan)
        ps.append(compute_psnr(pred, hr_hsi))
        ss.append(compute_sam_metric(pred, hr_hsi))
        es.append(compute_ergas(pred, hr_hsi, scale=SCALE))
        si.append(compute_ssim(pred.cpu(), hr_hsi.cpu()))
    avg = lambda l: sum(l) / max(len(l), 1)
    return avg(ps), avg(ss), avg(es), avg(si)


def train_model(variant, model, train_l, val_l, device, args, logger, best_overall):
    """Train one variant and return best metrics dict."""
    # V4 uses V4 loss; others use default CompositeLossV4 with V3-like weights
    if variant == 'v4':
        criterion = CompositeLossV4(
            lambda_l1=1.0, lambda_sam=0.10,
            lambda_edge=0.05, lambda_ssim=0.01, lambda_spec=0.05,
        ).to(device)
    else:
        criterion = CompositeLossV4(
            lambda_l1=1.0, lambda_sam=0.05,
            lambda_edge=0.05, lambda_ssim=0.05, lambda_spec=0.0,
        ).to(device)

    optimiser = optim.Adam(model.parameters(), lr=1e-3)
    scheduler = optim.lr_scheduler.CosineAnnealingLR(optimiser, T_max=args.epochs)

    ckpt_path = os.path.join(CKPT_DIR, f'{variant}_best.pth')
    best = {'psnr': -1e9, 'sam': 1e9, 'ergas': 1e9, 'epoch': 0}

    print(f"\n  Training {variant.upper()} for {args.epochs} epochs …")

    for epoch in range(1, args.epochs + 1):
        t0 = time.time()
        train_loss = train_one_epoch(model, train_l, criterion, optimiser, device)
        val_psnr, val_sam, val_ergas, val_ssim = validate(model, val_l, device)
        scheduler.step()
        t1 = time.time()

        is_best = val_psnr > best['psnr']
        if is_best:
            best.update({'psnr': val_psnr, 'sam': val_sam,
                         'ergas': val_ergas, 'ssim': val_ssim, 'epoch': epoch})
            torch.save(model.state_dict(), ckpt_path)

        lr_now = optimiser.param_groups[0]['lr']
        logger.log_epoch(variant, epoch, train_loss,
                         val_psnr, val_sam, val_ergas,
                         lr_now, t1-t0, is_best)

        star = ' *' if is_best else ''
        print(f"    Ep {epoch:3d}/{args.epochs} | "
              f"Loss={train_loss:.4f} | "
              f"PSNR={val_psnr:.3f} | "
              f"SAM={val_sam:.3f} | "
              f"ERGAS={val_ergas:.3f} | "
              f"SSIM={val_ssim:.4f}{star}")

    print(f"  Best {variant.upper()}: PSNR={best['psnr']:.3f} "
          f"SAM={best['sam']:.3f} ERGAS={best['ergas']:.3f} "
          f"SSIM={best.get('ssim',0):.4f} (epoch {best['epoch']})")
    return best


# ============================================================================
# Comparison Plot
# ============================================================================

def plot_comparison(history: dict):
    """6-panel comparison plot for all 4 variants."""
    if not HAS_MPL or not history:
        return
    fig, axes = plt.subplots(2, 3, figsize=(16, 8))
    colours = {'old': 'tab:blue', 'improved': 'tab:orange',
               'v3': 'tab:green', 'v4': 'tab:purple'}

    for variant, h in history.items():
        c = colours.get(variant, 'gray')
        epochs = list(range(1, len(h['train_loss']) + 1))
        axes[0, 0].plot(epochs, h['train_loss'], label=variant.upper(), color=c)
        axes[0, 1].plot(epochs, h['val_psnr'],  label=variant.upper(), color=c)
        axes[0, 2].plot(epochs, h['val_sam'],   label=variant.upper(), color=c)
        axes[1, 0].plot(epochs, h['val_ergas'], label=variant.upper(), color=c)

    for ax, title, ylabel in [
        (axes[0, 0], 'Training Loss',          'Loss'),
        (axes[0, 1], 'Validation PSNR (+)',    'PSNR (dB)'),
        (axes[0, 2], 'Validation SAM (-)',     'SAM (deg)'),
        (axes[1, 0], 'Validation ERGAS (-)',   'ERGAS'),
    ]:
        ax.set_title(title)
        ax.set_xlabel('Epoch')
        ax.set_ylabel(ylabel)
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

    # Parameter count bar chart
    ax_bar = axes[1, 1]
    if history:
        variant_labels = list(history.keys())
        param_counts   = [history[v].get('params', 0) / 1e6 for v in variant_labels]
        bar_colours    = [colours.get(v, 'gray') for v in variant_labels]
        ax_bar.bar(variant_labels, param_counts, color=bar_colours)
        ax_bar.set_title('Parameter Count')
        ax_bar.set_ylabel('Params (M)')
        for i, v in enumerate(param_counts):
            ax_bar.text(i, v + 0.01, f'{v:.2f}M', ha='center', fontsize=8)

    # Best PSNR bar chart
    ax_psnr = axes[1, 2]
    if history:
        best_psnrs = [history[v].get('best_psnr', 0) for v in variant_labels]
        ax_psnr.bar(variant_labels, best_psnrs, color=bar_colours)
        ax_psnr.set_title('Best Validation PSNR')
        ax_psnr.set_ylabel('PSNR (dB)')
        for i, v in enumerate(best_psnrs):
            ax_psnr.text(i, v + 0.05, f'{v:.2f}', ha='center', fontsize=8)

    plt.suptitle('VMamba Pansharpening — V1 vs V2 vs V3 vs V4', fontsize=13)
    plt.tight_layout()
    out = os.path.join(PLOTS_DIR, 'comparison_all.png')
    plt.savefig(out, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"\n  Comparison plot -> {out}")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--epochs',        default=5,    type=int)
    parser.add_argument('--batch_size',    default=1,    type=int)
    parser.add_argument('--d_model',       default=32,   type=int)
    parser.add_argument('--d_state',       default=8,    type=int)
    parser.add_argument('--patch_size',    default=32,   type=int)
    parser.add_argument('--skip_old',      action='store_true')
    parser.add_argument('--skip_improved', action='store_true')
    parser.add_argument('--skip_v3',       action='store_true')
    parser.add_argument('--skip_v4',       action='store_true')
    parser.add_argument('--save_img_every', default=5, type=int,
                        help='Save input/output images every N epochs (0=only final)')
    args = parser.parse_args()

    assert args.patch_size >= 32, \
        "patch_size must be ≥ 32 (4-stage backbone needs 4×4 bottleneck minimum)"

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"\nDevice : {device}")

    # Dataset
    print("\nLoading Chikusei …")
    train_l, val_l, test_l = create_dataloaders_overlap(
        dataset='chikusei', mat_path=CHIKUSEI_PATH,
        batch_size=args.batch_size, patch_size=args.patch_size,
        overlap=args.patch_size // 2, scale=SCALE, num_workers=0,
    )
    print(f"  train={len(train_l.dataset)}  val={len(val_l.dataset)}  "
          f"test={len(test_l.dataset)}")

    # Excel logger
    ts      = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    xl_path = os.path.join(RESULTS_DIR, f'training_log_{ts}.xlsx')
    logger  = ExcelLogger(xl_path)

    VARIANTS = []
    if not args.skip_old:      VARIANTS.append('old')
    if not args.skip_improved: VARIANTS.append('improved')
    if not args.skip_v3:       VARIANTS.append('v3')
    if not args.skip_v4:       VARIANTS.append('v4')

    if not VARIANTS:
        print("All variants skipped — nothing to do."); return

    # Build models
    models = {}
    for v in VARIANTS:
        m = create_vmamba_model(v, IN_CHANNELS, IN_CHANNELS,
                                scale=SCALE, d_model=args.d_model,
                                d_state=args.d_state).to(device)
        models[v] = m
        total, _ = count_parameters(m)
        print(f"  {v:<10}  params={total:>10,}")

    # ── Training loop with image saving ──────────────────────────────────────
    history = {v: {'train_loss': [], 'val_psnr': [], 'val_sam': [],
                   'val_ergas': [], 'params': count_parameters(models[v])[0]}
               for v in VARIANTS}
    best_all = {}

    # Grab one fixed batch for image visualisation
    vis_batch = next(iter(val_l))

    for v in VARIANTS:
        m = models[v]
        criterion = (CompositeLossV4(lambda_l1=1.0, lambda_sam=0.10,
                                     lambda_edge=0.05, lambda_ssim=0.01,
                                     lambda_spec=0.05)
                     if v == 'v4' else
                     CompositeLossV4(lambda_l1=1.0, lambda_sam=0.05,
                                     lambda_edge=0.05, lambda_ssim=0.05,
                                     lambda_spec=0.0)).to(device)
        optimiser = optim.Adam(m.parameters(), lr=1e-3)
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimiser, T_max=args.epochs)
        ckpt_path = os.path.join(CKPT_DIR, f'{v}_best.pth')
        best = {'psnr': -1e9, 'sam': 1e9, 'ergas': 1e9, 'epoch': 0}

        print(f"\n  Training {v.upper()} for {args.epochs} epochs …")

        for epoch in range(1, args.epochs + 1):
            t0 = time.time()
            train_loss = train_one_epoch(m, train_l, criterion, optimiser, device)
            val_psnr, val_sam, val_ergas, val_ssim = validate(m, val_l, device)
            scheduler.step()
            t1 = time.time()

            history[v]['train_loss'].append(train_loss)
            history[v]['val_psnr'].append(val_psnr)
            history[v]['val_sam'].append(val_sam)
            history[v]['val_ergas'].append(val_ergas)

            is_best = val_psnr > best['psnr']
            if is_best:
                best.update({'psnr': val_psnr, 'sam': val_sam,
                             'ergas': val_ergas, 'ssim': val_ssim, 'epoch': epoch})
                torch.save(m.state_dict(), ckpt_path)

            lr_now = optimiser.param_groups[0]['lr']
            logger.log_epoch(v, epoch, train_loss,
                             val_psnr, val_sam, val_ergas,
                             lr_now, t1-t0, is_best)

            star = ' *' if is_best else ''
            print(f"    Ep {epoch:3d}/{args.epochs} | "
                  f"Loss={train_loss:.4f} | "
                  f"PSNR={val_psnr:.3f} | "
                  f"SAM={val_sam:.3f} | "
                  f"ERGAS={val_ergas:.3f} | "
                  f"SSIM={val_ssim:.4f}{star}")

            # Save images periodically
            save_interval = args.save_img_every if args.save_img_every > 0 else args.epochs
            if epoch % save_interval == 0 or epoch == args.epochs:
                m.eval()
                with torch.no_grad():
                    vb_lr  = vis_batch['lr_hsi'].to(device)
                    vb_pan = vis_batch['hr_pan'].to(device)
                    pred   = m(vb_lr, vb_pan).cpu()
                save_images(
                    {k: vis_batch[k] for k in ['lr_hsi','hr_pan','hr_hsi']},
                    {v: pred},
                    epoch, tag=v,
                )
                m.train()

        history[v]['best_psnr']  = best['psnr']
        history[v]['best_sam']   = best['sam']
        history[v]['best_ergas'] = best['ergas']
        best['params'] = count_parameters(m)[0] / 1e6
        best_all[v] = best

        print(f"  Best {v.upper()}: PSNR={best['psnr']:.3f} "
              f"SAM={best['sam']:.3f} ERGAS={best['ergas']:.3f} "
              f"(epoch {best['epoch']})")

    # ── Final multi-model image comparison ───────────────────────────────────
    print("\n  Generating final multi-model image comparison …")
    final_outputs = {}
    for v in VARIANTS:
        ckpt = os.path.join(CKPT_DIR, f'{v}_best.pth')
        if os.path.isfile(ckpt):
            models[v].load_state_dict(torch.load(ckpt, map_location=device))
        models[v].eval()
        with torch.no_grad():
            vb_lr  = vis_batch['lr_hsi'].to(device)
            vb_pan = vis_batch['hr_pan'].to(device)
            final_outputs[v] = models[v](vb_lr, vb_pan).cpu()

    save_images(
        {k: vis_batch[k] for k in ['lr_hsi','hr_pan','hr_hsi']},
        final_outputs,
        epoch=args.epochs, tag='final_all',
    )
    save_spectral_curves(
        {k: vis_batch[k] for k in ['lr_hsi','hr_pan','hr_hsi']},
        final_outputs,
        epoch=args.epochs,
    )

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 68)
    print(f"{'Variant':<12} {'Best PSNR':>10} {'Best SAM':>10} "
          f"{'Best ERGAS':>12} {'Params (M)':>12}")
    print("=" * 68)
    for v in VARIANTS:
        b = best_all[v]
        total_p, _ = count_parameters(models[v])
        print(f"  {v:<10} {b['psnr']:>10.3f} {b['sam']:>10.3f} "
              f"{b['ergas']:>12.3f} {total_p/1e6:>12.3f}")
    print("=" * 68)

    # Write summary to excel
    logger.write_summary({v: {
        'psnr': best_all[v]['psnr'],
        'sam':  best_all[v]['sam'],
        'ergas': best_all[v]['ergas'],
        'params': count_parameters(models[v])[0] / 1e6,
    } for v in VARIANTS})

    # Save JSON results
    json_path = os.path.join(RESULTS_DIR, 'summary_all.json')
    with open(json_path, 'w') as f:
        json.dump({v: best_all[v] for v in VARIANTS}, f, indent=2)
    print(f"\n  JSON  -> {json_path}")
    if HAS_XL:
        print(f"  Excel -> {xl_path}")
    print(f"  Images -> {IMAGES_DIR}/")

    # Comparison plot
    plot_comparison(history)


if __name__ == '__main__':
    main()
