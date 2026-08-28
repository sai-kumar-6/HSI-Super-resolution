"""
json_to_excel.py
Converts v5_training_history.json to a formatted Excel file.
Output: version5/comparison/results/v5_training_history.xlsx
"""

import json, os
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as opnumbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

HERE    = os.path.dirname(os.path.abspath(__file__))
JSON    = os.path.join(HERE, 'version5', 'comparison', 'results', 'v5_training_history.json')
OUT     = os.path.join(HERE, 'version5', 'comparison', 'results', 'v5_training_history.xlsx')

with open(JSON) as f:
    history = json.load(f)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'V5 Training History'

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='1A237E')   # dark blue
HEADER_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
BEST_FILL    = PatternFill('solid', fgColor='E8F5E9')   # light green
BEST_FONT    = Font(name='Calibri', bold=True, color='1B5E20', size=10)
NORMAL_FONT  = Font(name='Calibri', size=10)
ALT_FILL     = PatternFill('solid', fgColor='F5F5F5')
CENTER       = Alignment(horizontal='center', vertical='center')
thin         = Side(style='thin', color='BDBDBD')
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Column headers ────────────────────────────────────────────────────────────
COLS = [
    ('Epoch',      'epoch',      '0',        False, None),
    ('Train Loss', 'train_loss', '0.000000', True,  'min'),
    ('Val PSNR',   'val_psnr',   '0.000000', True,  'max'),
    ('Val SAM',    'val_sam',    '0.000000', True,  'min'),
    ('Val ERGAS',  'val_ergas',  '0.000000', True,  'min'),
    ('Val SSIM',   'val_ssim',   '0.000000', True,  'max'),
]

# ── Write header row ──────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 22
for col_idx, (header, _, _, _, _) in enumerate(COLS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER

# ── Find best values per metric ───────────────────────────────────────────────
best = {}
for _, key, _, track, mode in COLS:
    if not track:
        continue
    vals = [h[key] for h in history]
    best[key] = min(vals) if mode == 'min' else max(vals)

# ── Write data rows ───────────────────────────────────────────────────────────
for row_idx, record in enumerate(history, start=2):
    ws.row_dimensions[row_idx].height = 16
    is_alt = (row_idx % 2 == 0)

    for col_idx, (_, key, fmt, track, mode) in enumerate(COLS, start=1):
        val  = record[key]
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.number_format = fmt
        cell.alignment     = CENTER
        cell.border        = BORDER

        is_best = track and (val == best[key])
        if is_best:
            cell.fill = BEST_FILL
            cell.font = BEST_FONT
        else:
            cell.fill = ALT_FILL if is_alt else PatternFill()
            cell.font = NORMAL_FONT

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [8, 14, 14, 12, 14, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Color-scale conditional formatting ───────────────────────────────────────
n = len(history) + 1   # last data row

# PSNR: low=red → high=green
ws.conditional_formatting.add(
    f'C2:C{n}',
    ColorScaleRule(start_type='min', start_color='FF5252',
                   mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                   end_type='max', end_color='69F0AE'))

# ERGAS: low=green → high=red
ws.conditional_formatting.add(
    f'E2:E{n}',
    ColorScaleRule(start_type='min', start_color='69F0AE',
                   mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                   end_type='max', end_color='FF5252'))

# SAM: low=green → high=red
ws.conditional_formatting.add(
    f'D2:D{n}',
    ColorScaleRule(start_type='min', start_color='69F0AE',
                   mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                   end_type='max', end_color='FF5252'))

# SSIM: low=red → high=green
ws.conditional_formatting.add(
    f'F2:F{n}',
    ColorScaleRule(start_type='min', start_color='FF5252',
                   mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                   end_type='max', end_color='69F0AE'))

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Summary')
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 10

summary_header = [('Metric', 'Best Value', 'Best Epoch')]
rows = [
    ('Train Loss (min)', best['train_loss'], next(h['epoch'] for h in history if h['train_loss'] == best['train_loss'])),
    ('Val PSNR (max)',   best['val_psnr'],   next(h['epoch'] for h in history if h['val_psnr']   == best['val_psnr'])),
    ('Val SAM (min)',    best['val_sam'],     next(h['epoch'] for h in history if h['val_sam']    == best['val_sam'])),
    ('Val ERGAS (min)',  best['val_ergas'],   next(h['epoch'] for h in history if h['val_ergas']  == best['val_ergas'])),
    ('Val SSIM (max)',   best['val_ssim'],    next(h['epoch'] for h in history if h['val_ssim']   == best['val_ssim'])),
]

for col_idx, hdr in enumerate(['Metric', 'Best Value', 'Best Epoch'], start=1):
    cell = ws2.cell(row=1, column=col_idx, value=hdr)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = BORDER

for r, (metric, val, ep) in enumerate(rows, start=2):
    for col_idx, v in enumerate([metric, val, ep], start=1):
        cell = ws2.cell(row=r, column=col_idx, value=v)
        cell.font      = Font(name='Calibri', bold=True, size=11)
        cell.fill      = BEST_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        if col_idx == 2:
            cell.number_format = '0.000000'

ws2.row_dimensions[1].height = 22

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f'Saved: {OUT}')
print(f'  Rows: {len(history)} epochs')
print(f'  Sheets: "V5 Training History", "Summary"')
