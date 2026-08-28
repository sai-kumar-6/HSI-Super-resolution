"""
compare_all.py  —  V3 Edition
M.Tech Thesis – Hyperspectral Pansharpening

Trains and compares all THREE VMamba variants on Chikusei:
  • Old VMamba     (V1: CrossAttention + Sobel + PixelShuffle)
  • Improved VMamba (V2: CrossMamba + LearnableHF + ResidualUpsample)
  • V3 VMamba      (V3: SDIM + MultiScaleSpectral + SpectralSSM)

Outputs
───────
  comparison/checkpoints/old_best.pth
  comparison/checkpoints/improved_best.pth
  comparison/checkpoints/v3_best.pth
  comparison/plots/comparison_all.png
  comparison/results/summary_all.json
  comparison/results/training_log_<timestamp>.xlsx   ← per-epoch Excel log

Usage
─────
  python comparison/compare_all.py
  python comparison/compare_all.py --epochs 30 --d_model 32 --patch_size 32
"""

import sys
import os
import json
import time
import argparse
import copy
from datetime import datetime

_HERE    = os.path.dirname(os.path.abspath(__file__))
_PROJECT = os.path.dirname(os.path.dirname(_HERE))   # Mtech project root
_V3      = os.path.dirname(_HERE)                    # version3 folder
sys.path.insert(0, _PROJECT)
sys.path.insert(0, _V3)

import torch
import torch.nn as nn

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    def tqdm(x, **kw): return x

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from dataset_loader_overlap import create_dataloaders_overlap as _create_dl
from baseline_models import create_vmamba_model, count_parameters
from loss_functions  import CompositeLoss, compute_psnr, compute_sam_metric, compute_ergas, compute_ssim


# ── Paths ────────────────────────────────────────────────────────────────────
CHIKUSEI_PATH = os.path.join(_PROJECT, 'chikusei', 'chikusei.mat')
CKPT_DIR      = os.path.join(_HERE, 'checkpoints')
PLOT_DIR      = os.path.join(_HERE, 'plots')
RESULT_DIR    = os.path.join(_HERE, 'results')
for _d in (CKPT_DIR, PLOT_DIR, RESULT_DIR):
    os.makedirs(_d, exist_ok=True)

VARIANTS = ['old', 'improved', 'v3']
LABELS   = {'old': 'Old VMamba', 'improved': 'Improved VMamba', 'v3': 'V3 VMamba'}
COLOURS  = {'old': 'steelblue', 'improved': 'darkorange', 'v3': 'mediumseagreen'}


# ── Excel Logger ─────────────────────────────────────────────────────────────

class ExcelLogger:
    """Per-epoch Excel log with Epoch Log + Summary sheets."""

    EPOCH_COLS = [
        ('Model',           18), ('Epoch',  7), ('Train Loss',    12),
        ('Val PSNR (dB)',   14), ('Val SAM (deg)', 14), ('Val ERGAS', 12),
        ('Learning Rate',   14), ('Epoch Time (s)', 14), ('Best?',  7),
    ]
    SUMMARY_COLS = [
        ('Metric', 22), ('Old VMamba', 16), ('Improved VMamba', 18),
        ('V3 VMamba', 14), ('Best', 12),
    ]

    def __init__(self, path):
        self.path = path
        if not HAS_OPENPYXL:
            return
        self.wb = openpyxl.Workbook()
        self.ws_epoch   = self.wb.active
        self.ws_epoch.title = 'Epoch Log'
        self.ws_summary = self.wb.create_sheet('Summary')
        self._write_header(self.ws_epoch,   self.EPOCH_COLS)
        self._write_header(self.ws_summary, self.SUMMARY_COLS)
        self.wb.save(path)
        print(f"[Excel] {path}")

    def _hstyle(self):
        return dict(
            font      = Font(bold=True, color='FFFFFF', size=11),
            fill      = PatternFill('solid', fgColor='1F497D'),
            alignment = Alignment(horizontal='center', vertical='center'),
            border    = Border(bottom=Side(style='medium'),
                               right =Side(style='thin', color='AAAAAA')),
        )

    def _write_header(self, ws, cols):
        ws.row_dimensions[1].height = 20
        for ci, (h, w) in enumerate(cols, 1):
            cell = ws.cell(1, ci, h)
            for attr, val in self._hstyle().items():
                setattr(cell, attr, val)
            ws.column_dimensions[get_column_letter(ci)].width = w

    _BG = {'old': 'DCE6F1', 'improved': 'FFF2CC', 'v3': 'E2EFDA'}

    def log_epoch(self, variant, epoch, train_loss, val_psnr,
                  val_sam, val_ergas, lr, elapsed, is_best):
        if not HAS_OPENPYXL:
            return
        row  = self.ws_epoch.max_row + 1
        fill = PatternFill('solid', fgColor=self._BG.get(variant, 'F2F2F2'))
        vals = [LABELS[variant], epoch, round(train_loss, 6),
                round(val_psnr, 4), round(val_sam, 4), round(val_ergas, 4),
                lr, round(elapsed, 2), '*' if is_best else '']
        for ci, v in enumerate(vals, 1):
            cell = self.ws_epoch.cell(row, ci, v)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            if ci == 9 and is_best:
                cell.font = Font(bold=True, color='FF0000')
        self.wb.save(self.path)

    def write_summary(self, results: dict, t_ms: dict):
        if not HAS_OPENPYXL:
            return
        ws     = self.ws_summary
        green  = PatternFill('solid', fgColor='C6EFCE')
        red    = PatternFill('solid', fgColor='FFC7CE')
        metrics = [
            ('PSNR (dB)',       'best_psnr',  True ),
            ('SAM (deg)',       'best_sam',   False),
            ('ERGAS',          'best_ergas', False),
            ('Inference (ms)', None,         False),
            ('Params (M)',     None,         False),
        ]
        for ri, (name, key, hb) in enumerate(metrics, 2):
            if key:
                vals = {v: results[v][key] for v in VARIANTS}
            elif name.startswith('Inference'):
                vals = t_ms
            else:
                vals = {v: results[v]['n_params']/1e6 for v in VARIANTS}

            best_v = max(vals, key=vals.get) if hb else min(vals, key=vals.get)
            ws.cell(ri, 1, name).font = Font(bold=True)
            for ci, v in enumerate(VARIANTS, 2):
                cell = ws.cell(ri, ci, round(vals[v], 4))
                cell.alignment = Alignment(horizontal='center')
                cell.fill = green if v == best_v else red
            ws.cell(ri, 5, LABELS[best_v]).alignment = Alignment(horizontal='center')

        ws.cell(len(metrics)+3, 1, 'Generated').font = Font(italic=True, color='888888')
        ws.cell(len(metrics)+3, 2,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ).font = Font(italic=True, color='888888')
        self.wb.save(self.path)
        print(f"[Excel] Summary written -> {self.path}")


# ── Training helpers ─────────────────────────────────────────────────────────

def train_epoch(model, loader, optimizer, criterion, device, epoch, label):
    model.train()
    total, n = 0.0, 0
    pbar = tqdm(loader, desc=f"[{label}] Train E{epoch}", leave=False) if HAS_TQDM else loader
    for batch in pbar:
        lr_hsi = batch['lr_hsi'].to(device)
        hr_pan = batch['hr_pan'].to(device)
        hr_hsi = batch['hr_hsi'].to(device)
        optimizer.zero_grad()
        pred = model(lr_hsi, hr_pan)
        loss = criterion(pred, hr_hsi)
        loss.backward()
        nn.utils.clip_grad_norm_(model.parameters(), 1.0)
        optimizer.step()
        total += loss.item(); n += 1
        if HAS_TQDM:
            pbar.set_postfix({'loss': f'{loss.item():.4f}'})
    return total / max(n, 1)


@torch.no_grad()
def validate_epoch(model, loader, device, label, epoch):
    model.eval()
    ps, ss, es, si = [], [], [], []
    pbar = tqdm(loader, desc=f"[{label}] Val   E{epoch}", leave=False) if HAS_TQDM else loader
    for batch in pbar:
        lr_hsi = batch['lr_hsi'].to(device)
        hr_pan = batch['hr_pan'].to(device)
        hr_hsi = batch['hr_hsi'].to(device)
        pred = model(lr_hsi, hr_pan)
        ps.append(compute_psnr(pred, hr_hsi))
        ss.append(compute_sam_metric(pred, hr_hsi))
        es.append(compute_ergas(pred, hr_hsi))
        si.append(compute_ssim(pred.cpu(), hr_hsi.cpu()))
    avg = lambda lst: sum(lst) / max(len(lst), 1)
    return {'psnr': avg(ps), 'sam': avg(ss), 'ergas': avg(es), 'ssim': avg(si)}


def train_model(variant, cfg, train_loader, val_loader, device, logger=None):
    label = LABELS[variant]
    print(f"\n{'='*60}\n  Training {label}  ({cfg.epochs} epochs)\n{'='*60}")

    model = create_vmamba_model(
        variant=variant, in_ch=128, out_ch=128,
        scale=cfg.scale, d_model=cfg.d_model,
        d_state=cfg.d_state, num_blocks=cfg.num_blocks,
    ).to(device)

    total, _ = count_parameters(model)
    print(f"  Parameters: {total:,}\n")

    criterion = CompositeLoss(lambda_l1=1.0, lambda_sam=0.05,
                              lambda_edge=0.05, lambda_ssim=0.05)
    optimizer = torch.optim.AdamW(model.parameters(), lr=1e-4, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer, T_max=cfg.epochs, eta_min=1e-6)

    history   = {'train_losses': [], 'val_psnrs': [], 'val_sams': [], 'val_ergases': []}
    best_psnr = -float('inf')
    best_metrics = {}
    ckpt_path = os.path.join(CKPT_DIR, f'{variant}_best.pth')

    for epoch in range(1, cfg.epochs + 1):
        t0         = time.time()
        train_loss = train_epoch(model, train_loader, optimizer, criterion, device, epoch, label)
        val_met    = validate_epoch(model, val_loader, device, label, epoch)
        scheduler.step()
        elapsed    = time.time() - t0

        history['train_losses'].append(train_loss)
        history['val_psnrs'].append(val_met['psnr'])
        history['val_sams'].append(val_met['sam'])
        history['val_ergases'].append(val_met['ergas'])

        is_best = val_met['psnr'] > best_psnr
        if is_best:
            best_psnr    = val_met['psnr']
            best_metrics = val_met.copy()
            torch.save({'model_state_dict': copy.deepcopy(model.state_dict()),
                        'epoch': epoch, 'metrics': best_metrics}, ckpt_path)

        cur_lr = optimizer.param_groups[0]['lr']
        print(f"  E{epoch:3d}/{cfg.epochs}  loss={train_loss:.4f}  "
              f"PSNR={val_met['psnr']:.2f}  SAM={val_met['sam']:.3f}  "
              f"ERGAS={val_met['ergas']:.3f}  SSIM={val_met['ssim']:.4f}  "
              f"({'*BEST* ' if is_best else ''}{elapsed:.1f}s)")

        if logger:
            logger.log_epoch(variant, epoch, train_loss,
                             val_met['psnr'], val_met['sam'], val_met['ergas'],
                             cur_lr, elapsed, is_best)

    print(f"\n  Best PSNR = {best_psnr:.2f} dB  ->  {ckpt_path}")
    return {**history,
            'best_psnr':  best_metrics.get('psnr',  best_psnr),
            'best_sam':   best_metrics.get('sam',   0.0),
            'best_ergas': best_metrics.get('ergas', 0.0),
            'ckpt_path':  ckpt_path,
            'n_params':   total}


@torch.no_grad()
def measure_inference(model, device, scale=4, n_runs=20):
    model.eval()
    lr  = torch.randn(1, 128, 8, 8).to(device)
    pan = torch.randn(1, 1, 32, 32).to(device)
    for _ in range(3):
        model(lr, pan)   # warm-up
    times = []
    for _ in range(n_runs):
        t0 = time.perf_counter()
        model(lr, pan)
        times.append((time.perf_counter() - t0) * 1000)
    return sum(times) / len(times)


# ── Plot ─────────────────────────────────────────────────────────────────────

def generate_plots(results: dict, cfg):
    epochs = list(range(1, cfg.epochs + 1))
    fig, axes = plt.subplots(2, 3, figsize=(18, 10))
    fig.suptitle('VMamba Pansharpening — V1 vs V2 vs V3', fontsize=16, fontweight='bold')

    panels = [
        (axes[0,0], 'train_losses', 'Training Loss',        'Loss',       False),
        (axes[0,1], 'val_psnrs',    'Validation PSNR (dB)', 'PSNR (dB)',  False),
        (axes[0,2], 'val_sams',     'Validation SAM (°)',   'SAM (°)',    False),
    ]
    for ax, key, title, ylabel, _ in panels:
        for v in VARIANTS:
            ax.plot(epochs, results[v][key], label=LABELS[v],
                    color=COLOURS[v], linewidth=1.8)
        ax.set_title(title); ax.set_xlabel('Epoch'); ax.set_ylabel(ylabel)
        ax.legend(); ax.grid(True, alpha=0.3)

    # Best metrics bar
    ax = axes[1, 0]
    metric_keys = [('PSNR (dB)', 'best_psnr'), ('SAM (°)', 'best_sam'), ('ERGAS', 'best_ergas')]
    x = range(len(metric_keys)); w = 0.25
    for i, v in enumerate(VARIANTS):
        vals = [results[v][k] for _, k in metric_keys]
        bars = ax.bar([xi + i*w - w for xi in x], vals, w,
                      label=LABELS[v], color=COLOURS[v])
        for bar in bars:
            ax.text(bar.get_x() + bar.get_width()/2,
                    bar.get_height() + 0.01,
                    f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=7)
    ax.set_title('Best Metrics'); ax.set_xticks(list(x))
    ax.set_xticklabels([m for m, _ in metric_keys]); ax.legend(); ax.grid(True, alpha=0.3, axis='y')

    # Params
    ax = axes[1, 1]
    p_vals = [results[v]['n_params'] / 1e6 for v in VARIANTS]
    bars   = ax.bar(list(VARIANTS), p_vals, color=[COLOURS[v] for v in VARIANTS], width=0.4)
    ax.set_title('Parameter Count (M)'); ax.set_ylabel('Parameters (M)'); ax.grid(True, alpha=0.3, axis='y')
    for bar, val in zip(bars, p_vals):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02,
                f'{val:.2f}M', ha='center', va='bottom', fontsize=9)
    ax.set_xticklabels([LABELS[v] for v in VARIANTS], rotation=10, ha='right')

    # Inference time
    ax = axes[1, 2]
    if 'inference_ms' in results.get('old', {}):
        t_vals = [results[v].get('inference_ms', 0) for v in VARIANTS]
        bars   = ax.bar(list(VARIANTS), t_vals, color=[COLOURS[v] for v in VARIANTS], width=0.4)
        ax.set_title('Inference Time (ms / image)'); ax.set_ylabel('ms'); ax.grid(True, alpha=0.3, axis='y')
        for bar, val in zip(bars, t_vals):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                    f'{val:.1f}ms', ha='center', va='bottom', fontsize=9)
        ax.set_xticklabels([LABELS[v] for v in VARIANTS], rotation=10, ha='right')

    plt.tight_layout()
    out = os.path.join(PLOT_DIR, 'comparison_all.png')
    plt.savefig(out, dpi=150, bbox_inches='tight'); plt.close()
    print(f"  Plot -> {out}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--epochs',     default=5,   type=int)
    parser.add_argument('--batch_size', default=1,   type=int)
    parser.add_argument('--d_model',    default=32,  type=int)
    parser.add_argument('--d_state',    default=8,   type=int)
    parser.add_argument('--scale',      default=4,   type=int)
    parser.add_argument('--patch_size', default=32,  type=int,
                        help='HR patch size (min 32 for backbone)')
    parser.add_argument('--skip_old',      action='store_true')
    parser.add_argument('--skip_improved', action='store_true')
    parser.add_argument('--skip_v3',       action='store_true')
    args = parser.parse_args()

    args.num_blocks = [1, 1, 1, 1] if args.d_model <= 32 else [2, 2, 2, 2]
    args.patch_size = max(args.patch_size, 32)  # minimum 32 for 4-stage backbone

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"\nDevice     : {device}")
    print(f"Dataset    : {CHIKUSEI_PATH}")
    print(f"Epochs     : {args.epochs}")
    print(f"Patch size : {args.patch_size}x{args.patch_size}  ->  LR={args.patch_size//args.scale}x{args.patch_size//args.scale}")

    # Dataloaders
    print("\nLoading dataset …")
    train_l, val_l, _test_l = _create_dl(
        dataset='chikusei', mat_path=CHIKUSEI_PATH,
        batch_size=args.batch_size, patch_size=args.patch_size,
        overlap=args.patch_size // 2, scale=args.scale, num_workers=0,
    )
    print(f"  Train={len(train_l)}  Val={len(val_l)} batches")

    # Excel logger
    ts     = datetime.now().strftime('%Y%m%d_%H%M%S')
    logger = ExcelLogger(os.path.join(RESULT_DIR, f'training_log_{ts}.xlsx')) \
             if HAS_OPENPYXL else None

    # Train each variant
    results = {}
    skip = {'old': args.skip_old, 'improved': args.skip_improved, 'v3': args.skip_v3}
    for v in VARIANTS:
        if skip[v]:
            print(f"\n[SKIP] {LABELS[v]}")
            results[v] = {'train_losses': [], 'val_psnrs': [], 'val_sams': [],
                          'val_ergases': [], 'best_psnr': 0, 'best_sam': 99,
                          'best_ergas': 99, 'n_params': 0}
        else:
            results[v] = train_model(v, args, train_l, val_l, device, logger)

    # Inference benchmark
    print("\nBenchmarking inference …")
    t_ms = {}
    for v in VARIANTS:
        m = create_vmamba_model(v, 128, 128, scale=args.scale,
                                d_model=args.d_model, d_state=args.d_state,
                                num_blocks=args.num_blocks).to(device)
        ckpt = os.path.join(CKPT_DIR, f'{v}_best.pth')
        if os.path.isfile(ckpt):
            st = torch.load(ckpt, map_location=device)
            if isinstance(st, dict) and 'model_state_dict' in st:
                st = st['model_state_dict']
            m.load_state_dict(st)
        t_ms[v] = measure_inference(m, device)
        results[v]['inference_ms'] = t_ms[v]
        print(f"  {LABELS[v]:<20}: {t_ms[v]:.1f} ms")

    if logger:
        logger.write_summary(results, t_ms)

    generate_plots(results, args)

    # JSON summary
    serialisable = {
        'config': vars(args),
        'results': {
            v: {k: val for k, val in results[v].items()
                if k not in ('_model',)}
            for v in VARIANTS
        },
    }
    save_path = os.path.join(RESULT_DIR, 'summary_all.json')
    with open(save_path, 'w') as f:
        json.dump(serialisable, f, indent=2, default=str)
    print(f"  JSON -> {save_path}")

    # Final table
    print(f"\n{'='*65}")
    print(f"  FINAL COMPARISON  (Chikusei, scale={args.scale})")
    print(f"{'='*65}")
    print(f"  {'Metric':<20} {'Old':>10} {'Improved':>12} {'V3':>10}")
    print(f"  {'-'*53}")
    for metric, key in [('PSNR (dB)', 'best_psnr'), ('SAM (°)', 'best_sam'), ('ERGAS', 'best_ergas')]:
        vals = [results[v][key] for v in VARIANTS]
        print(f"  {metric:<20} {vals[0]:>10.3f} {vals[1]:>12.3f} {vals[2]:>10.3f}")
    params = [results[v]['n_params']/1e6 for v in VARIANTS]
    print(f"  {'Params (M)':<20} {params[0]:>10.2f} {params[1]:>12.2f} {params[2]:>10.2f}")
    t_vals = [t_ms[v] for v in VARIANTS]
    print(f"  {'Inference (ms)':<20} {t_vals[0]:>10.1f} {t_vals[1]:>12.1f} {t_vals[2]:>10.1f}")
    print(f"{'='*65}\n")


if __name__ == '__main__':
    import copy
    main()
