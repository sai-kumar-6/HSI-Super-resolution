"""
compare_vmamba.py
M.Tech Thesis – Hyperspectral Pansharpening

Trains BOTH VMamba variants (Old and Improved) on the Chikusei dataset
using Wald's protocol, then generates publication-quality comparison plots.

Usage
-----
    python comparison/compare_vmamba.py                      # defaults
    python comparison/compare_vmamba.py --epochs 50          # more epochs
    python comparison/compare_vmamba.py --batch_size 4 --d_model 32
"""

import sys
import os
import json
import time
import argparse
import copy
from datetime import datetime

# Make the project root importable from this subfolder
_HERE    = os.path.dirname(os.path.abspath(__file__))
_PROJECT = os.path.dirname(_HERE)
sys.path.insert(0, _PROJECT)

import torch
import torch.nn as nn
from torch.utils.data import DataLoader

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    def tqdm(x, **kw):
        return x

import matplotlib
matplotlib.use('Agg')          # non-interactive backend for server/script use
import matplotlib.pyplot as plt

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not found. Install with: pip install openpyxl")

# Try overlap dataloader first, fall back to standard one
try:
    from dataset_loader_overlap import create_dataloaders_overlap as _create_dl
    def create_dataloaders(mat_path, batch_size, patch_size, scale, num_workers=0):
        # Returns (train, val, test) – we only need train and val
        train_l, val_l, _test_l = _create_dl(
            dataset     = 'chikusei',
            mat_path    = mat_path,
            batch_size  = batch_size,
            patch_size  = patch_size,
            overlap     = patch_size // 2,
            scale       = scale,
            num_workers = num_workers,
        )
        return train_l, val_l
    print("[INFO] Using dataset_loader_overlap")
except ImportError:
    from dataset_loader import create_dataloaders as _create_dl_std
    def create_dataloaders(mat_path, batch_size, patch_size, scale, num_workers=0):
        return _create_dl_std(
            dataset     = 'chikusei',
            mat_path    = mat_path,
            batch_size  = batch_size,
            patch_size  = patch_size,
            scale       = scale,
            num_workers = num_workers,
        )
    print("[INFO] Using dataset_loader (standard)")

from baseline_models import create_vmamba_model, count_parameters
from loss_functions  import (
    CompositeLoss,
    compute_psnr,
    compute_sam_metric,
    compute_ergas,
)


# ============================================================================
# Paths
# ============================================================================

CHIKUSEI_PATH = os.path.join(_PROJECT, 'chikusei', 'chikusei.mat')
CKPT_DIR      = os.path.join(_HERE, 'checkpoints')
PLOT_DIR      = os.path.join(_HERE, 'plots')
RESULT_DIR    = os.path.join(_HERE, 'results')

for _d in (CKPT_DIR, PLOT_DIR, RESULT_DIR):
    os.makedirs(_d, exist_ok=True)


# ============================================================================
# Excel Logger
# ============================================================================

class ExcelLogger:
    """
    Writes per-epoch training/validation metrics to an Excel workbook.
    Creates two sheets:
      - 'Epoch Log'  : one row per epoch per model (train + val metrics)
      - 'Summary'    : best metrics side-by-side comparison
    """

    # Column definitions: (header, width)
    EPOCH_COLS = [
        ('Model',            18),
        ('Epoch',             7),
        ('Train Loss',       12),
        ('Val PSNR (dB)',    14),
        ('Val SAM (deg)',    14),
        ('Val ERGAS',        12),
        ('Learning Rate',    14),
        ('Epoch Time (s)',   14),
        ('Best?',             7),
    ]

    SUMMARY_COLS = [
        ('Metric',           22),
        ('Old VMamba',       16),
        ('Improved VMamba',  18),
        ('Delta',            12),
        ('Better',           10),
    ]

    def __init__(self, path: str):
        self.path = path
        if not HAS_OPENPYXL:
            return

        self.wb = openpyxl.Workbook()

        # ---- Epoch Log sheet ----
        self.ws_epoch = self.wb.active
        self.ws_epoch.title = 'Epoch Log'
        self._write_epoch_header()

        # ---- Summary sheet ----
        self.ws_summary = self.wb.create_sheet('Summary')
        self._write_summary_header()

        self.wb.save(self.path)
        print(f"[Excel] Log file created → {self.path}")

    def _header_style(self):
        return {
            'font':      Font(bold=True, color='FFFFFF', size=11),
            'fill':      PatternFill('solid', fgColor='1F497D'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border':    Border(
                bottom=Side(style='medium', color='000000'),
                right =Side(style='thin',   color='AAAAAA'),
            ),
        }

    def _apply_style(self, cell, style: dict):
        for attr, val in style.items():
            setattr(cell, attr, val)

    def _write_epoch_header(self):
        style = self._header_style()
        self.ws_epoch.row_dimensions[1].height = 20
        for col_idx, (header, width) in enumerate(self.EPOCH_COLS, start=1):
            cell = self.ws_epoch.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, style)
            self.ws_epoch.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_summary_header(self):
        style = self._header_style()
        self.ws_summary.row_dimensions[1].height = 20
        for col_idx, (header, width) in enumerate(self.SUMMARY_COLS, start=1):
            cell = self.ws_summary.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, style)
            self.ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    def log_epoch(self, model_label: str, epoch: int, total_epochs: int,
                  train_loss: float, val_psnr: float, val_sam: float,
                  val_ergas: float, lr: float, elapsed: float, is_best: bool):
        """Append one epoch row and save."""
        if not HAS_OPENPYXL:
            return

        row = self.ws_epoch.max_row + 1

        # Alternating row colours by model
        bg = 'DCE6F1' if 'Old' in model_label else 'E2EFDA'
        fill = PatternFill('solid', fgColor=bg)

        values = [
            model_label,
            epoch,
            round(train_loss, 6),
            round(val_psnr,   4),
            round(val_sam,    4),
            round(val_ergas,  4),
            lr,
            round(elapsed,    2),
            '*' if is_best else '',
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = self.ws_epoch.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            # Bold the 'Best?' column marker
            if col_idx == 9 and is_best:
                cell.font = Font(bold=True, color='FF0000')

        self.wb.save(self.path)

    def write_summary(self, hist_old: dict, hist_imp: dict,
                      t_old_ms: float, t_imp_ms: float):
        """Write the Summary sheet with best metrics comparison."""
        if not HAS_OPENPYXL:
            return

        ws = self.ws_summary
        metrics = [
            ('PSNR (dB)',       hist_old['best_psnr'],  hist_imp['best_psnr'],  True),
            ('SAM (deg)',       hist_old['best_sam'],   hist_imp['best_sam'],   False),
            ('ERGAS',          hist_old['best_ergas'], hist_imp['best_ergas'], False),
            ('Inference (ms)', t_old_ms,               t_imp_ms,              False),
            ('Params (M)',     hist_old['n_params']/1e6, hist_imp['n_params']/1e6, False),
        ]

        green_fill = PatternFill('solid', fgColor='C6EFCE')
        red_fill   = PatternFill('solid', fgColor='FFC7CE')

        for r_idx, (name, old_v, imp_v, higher_better) in enumerate(metrics, start=2):
            delta = imp_v - old_v
            imp_better = (delta > 0) == higher_better

            ws.cell(row=r_idx, column=1, value=name).font = Font(bold=True)

            c_old = ws.cell(row=r_idx, column=2, value=round(old_v, 4))
            c_imp = ws.cell(row=r_idx, column=3, value=round(imp_v, 4))
            c_dlt = ws.cell(row=r_idx, column=4, value=round(delta, 4))
            c_btr = ws.cell(row=r_idx, column=5, value='Improved ↑' if imp_better else 'No change' if delta == 0 else 'Old ↑')

            # Highlight the better value
            for cell in [c_old, c_imp]:
                cell.alignment = Alignment(horizontal='center')
            c_imp.fill = green_fill if imp_better else red_fill
            c_dlt.alignment = Alignment(horizontal='center')
            c_btr.alignment = Alignment(horizontal='center')

        # Config info
        ws.cell(row=len(metrics) + 3, column=1, value='Generated').font = Font(italic=True, color='888888')
        ws.cell(row=len(metrics) + 3, column=2,
                value=datetime.now().strftime('%Y-%m-%d %H:%M:%S')).font = Font(italic=True, color='888888')

        self.wb.save(self.path)
        print(f"[Excel] Summary sheet written → {self.path}")


# ============================================================================
# Training loop for one epoch
# ============================================================================

def train_epoch(model, loader, optimizer, criterion, device, epoch, label):
    model.train()
    total_loss = 0.0
    n_batches  = 0

    pbar = tqdm(loader, desc=f"[{label}] Train E{epoch}", leave=False) if HAS_TQDM else loader
    for batch in pbar:
        lr_hsi = batch['lr_hsi'].to(device)
        hr_pan = batch['hr_pan'].to(device)
        hr_hsi = batch['hr_hsi'].to(device)

        optimizer.zero_grad()
        pred = model(lr_hsi, hr_pan)
        loss = criterion(pred, hr_hsi)
        loss.backward()

        # Gradient clipping for stability
        nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
        optimizer.step()

        total_loss += loss.item()
        n_batches  += 1

        if HAS_TQDM:
            pbar.set_postfix({'loss': f'{loss.item():.4f}'})

    return total_loss / max(n_batches, 1)


# ============================================================================
# Validation loop
# ============================================================================

@torch.no_grad()
def validate_epoch(model, loader, device, label, epoch):
    model.eval()
    psnr_list, sam_list, ergas_list = [], [], []

    pbar = tqdm(loader, desc=f"[{label}]   Val  E{epoch}", leave=False) if HAS_TQDM else loader
    for batch in pbar:
        lr_hsi = batch['lr_hsi'].to(device)
        hr_pan = batch['hr_pan'].to(device)
        hr_hsi = batch['hr_hsi'].to(device)

        pred = model(lr_hsi, hr_pan)
        psnr_list.append( compute_psnr(pred, hr_hsi) )
        sam_list.append(  compute_sam_metric(pred, hr_hsi) )
        ergas_list.append(compute_ergas(pred, hr_hsi) )

    avg = lambda lst: sum(lst) / max(len(lst), 1)
    return {
        'psnr':  avg(psnr_list),
        'sam':   avg(sam_list),
        'ergas': avg(ergas_list),
    }


# ============================================================================
# Full training run for one model variant
# ============================================================================

def train_model(
    variant:    str,
    cfg:        argparse.Namespace,
    train_loader,
    val_loader,
    device:     torch.device,
    logger:     'ExcelLogger | None' = None,
) -> dict:
    """
    Train one model variant and return history + best metrics.

    Returns
    -------
    dict with keys: train_losses, val_psnrs, val_sams, val_ergases,
                    best_psnr, best_sam, best_ergas, ckpt_path
    """
    label = 'Old VMamba' if variant == 'old' else 'Improved VMamba'
    print(f"\n{'='*60}")
    print(f"  Training {label}  ({cfg.epochs} epochs)")
    print(f"{'='*60}")

    model = create_vmamba_model(
        variant    = variant,
        in_ch      = 128,
        out_ch     = 128,
        scale      = cfg.scale,
        d_model    = cfg.d_model,
        num_blocks = cfg.num_blocks,
    ).to(device)

    total, _ = count_parameters(model)
    print(f"  Parameters: {total:,}\n")

    criterion = CompositeLoss(lambda_l1=1.0, lambda_sam=0.05,
                              lambda_edge=0.05, lambda_ssim=0.05)
    optimizer = torch.optim.AdamW(model.parameters(), lr=1e-4, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer, T_max=cfg.epochs, eta_min=1e-6
    )

    history = {
        'train_losses': [],
        'val_psnrs':    [],
        'val_sams':     [],
        'val_ergases':  [],
    }
    best_psnr      = -float('inf')
    best_state     = None
    best_metrics   = {}
    ckpt_path      = os.path.join(CKPT_DIR, f'{variant}_best.pth')

    for epoch in range(1, cfg.epochs + 1):
        t0         = time.time()
        train_loss = train_epoch(model, train_loader, optimizer, criterion,
                                 device, epoch, label)
        val_met    = validate_epoch(model, val_loader, device, label, epoch)
        scheduler.step()
        elapsed    = time.time() - t0

        history['train_losses'].append(train_loss)
        history['val_psnrs'].append(   val_met['psnr'])
        history['val_sams'].append(    val_met['sam'])
        history['val_ergases'].append( val_met['ergas'])

        # Save best checkpoint
        is_best = val_met['psnr'] > best_psnr
        if is_best:
            best_psnr    = val_met['psnr']
            best_state   = copy.deepcopy(model.state_dict())
            best_metrics = val_met.copy()
            torch.save({'model_state_dict': best_state,
                        'epoch': epoch,
                        'metrics': best_metrics},
                       ckpt_path)

        current_lr = optimizer.param_groups[0]['lr']

        print(
            f"  E{epoch:3d}/{cfg.epochs}  "
            f"loss={train_loss:.4f}  "
            f"PSNR={val_met['psnr']:.2f} dB  "
            f"SAM={val_met['sam']:.3f}°  "
            f"ERGAS={val_met['ergas']:.3f}  "
            f"({'*BEST* ' if is_best else ''}{elapsed:.1f}s)"
        )

        # Log to Excel after every epoch
        if logger is not None:
            logger.log_epoch(
                model_label = label,
                epoch       = epoch,
                total_epochs= cfg.epochs,
                train_loss  = train_loss,
                val_psnr    = val_met['psnr'],
                val_sam     = val_met['sam'],
                val_ergas   = val_met['ergas'],
                lr          = current_lr,
                elapsed     = elapsed,
                is_best     = is_best,
            )

    print(f"\n  Best PSNR = {best_psnr:.2f} dB  →  {ckpt_path}")

    return {
        **history,
        'best_psnr':  best_metrics.get('psnr',  best_psnr),
        'best_sam':   best_metrics.get('sam',   0.0),
        'best_ergas': best_metrics.get('ergas', 0.0),
        'ckpt_path':  ckpt_path,
        'n_params':   total,
    }


# ============================================================================
# Inference-time benchmark
# ============================================================================

@torch.no_grad()
def measure_inference_time(model, device, batch_size=1, n_runs=20,
                            in_ch=128, scale=4, lr_h=16, lr_w=16):
    model.eval()
    lr  = torch.randn(batch_size, in_ch, lr_h, lr_w).to(device)
    pan = torch.randn(batch_size, 1, lr_h * scale, lr_w * scale).to(device)

    # Warm up
    for _ in range(3):
        _ = model(lr, pan)

    times = []
    for _ in range(n_runs):
        if device.type == 'cuda':
            torch.cuda.synchronize()
        t0 = time.perf_counter()
        _ = model(lr, pan)
        if device.type == 'cuda':
            torch.cuda.synchronize()
        times.append((time.perf_counter() - t0) * 1000)

    return sum(times) / len(times)   # ms


# ============================================================================
# Plot generation
# ============================================================================

def generate_plots(hist_old: dict, hist_imp: dict, cfg: argparse.Namespace):
    """Create and save all comparison plots."""
    epochs = list(range(1, cfg.epochs + 1))
    fig, axes = plt.subplots(2, 3, figsize=(18, 10))
    fig.suptitle('VMamba Pansharpening – Old vs Improved', fontsize=16, fontweight='bold')

    # ---- 1. Training loss ----
    ax = axes[0, 0]
    ax.plot(epochs, hist_old['train_losses'], label='Old VMamba',      color='steelblue',  linewidth=1.8)
    ax.plot(epochs, hist_imp['train_losses'], label='Improved VMamba', color='darkorange', linewidth=1.8)
    ax.set_title('Training Loss')
    ax.set_xlabel('Epoch')
    ax.set_ylabel('Loss')
    ax.legend()
    ax.grid(True, alpha=0.3)

    # ---- 2. Validation PSNR ----
    ax = axes[0, 1]
    ax.plot(epochs, hist_old['val_psnrs'], label='Old VMamba',      color='steelblue',  linewidth=1.8)
    ax.plot(epochs, hist_imp['val_psnrs'], label='Improved VMamba', color='darkorange', linewidth=1.8)
    ax.set_title('Validation PSNR (dB)')
    ax.set_xlabel('Epoch')
    ax.set_ylabel('PSNR (dB)')
    ax.legend()
    ax.grid(True, alpha=0.3)

    # ---- 3. Validation SAM ----
    ax = axes[0, 2]
    ax.plot(epochs, hist_old['val_sams'], label='Old VMamba',      color='steelblue',  linewidth=1.8)
    ax.plot(epochs, hist_imp['val_sams'], label='Improved VMamba', color='darkorange', linewidth=1.8)
    ax.set_title('Validation SAM (degrees, lower=better)')
    ax.set_xlabel('Epoch')
    ax.set_ylabel('SAM (°)')
    ax.legend()
    ax.grid(True, alpha=0.3)

    # ---- 4. Best metrics bar chart ----
    ax     = axes[1, 0]
    labels = ['PSNR (dB)', 'SAM (°)', 'ERGAS']
    old_v  = [hist_old['best_psnr'], hist_old['best_sam'], hist_old['best_ergas']]
    imp_v  = [hist_imp['best_psnr'], hist_imp['best_sam'], hist_imp['best_ergas']]
    x      = range(len(labels))
    w      = 0.35
    bars1  = ax.bar([i - w/2 for i in x], old_v, w, label='Old VMamba',      color='steelblue')
    bars2  = ax.bar([i + w/2 for i in x], imp_v, w, label='Improved VMamba', color='darkorange')
    ax.set_title('Best Metrics Comparison')
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels)
    ax.legend()
    ax.grid(True, alpha=0.3, axis='y')
    for bar in list(bars1) + list(bars2):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.01 * max(max(old_v), max(imp_v)),
                f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=8)

    # ---- 5. Parameter count ----
    ax       = axes[1, 1]
    p_labels = ['Old VMamba', 'Improved VMamba']
    p_vals   = [hist_old['n_params'] / 1e6, hist_imp['n_params'] / 1e6]
    colours  = ['steelblue', 'darkorange']
    bars     = ax.bar(p_labels, p_vals, color=colours, width=0.4)
    ax.set_title('Parameter Count (M)')
    ax.set_ylabel('Parameters (Millions)')
    ax.grid(True, alpha=0.3, axis='y')
    for bar, val in zip(bars, p_vals):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.02 * max(p_vals),
                f'{val:.2f}M', ha='center', va='bottom', fontsize=10)

    # ---- 6. Inference time (computed live if available) ----
    ax = axes[1, 2]
    if 'inference_ms_old' in hist_old and 'inference_ms_imp' in hist_imp:
        t_vals = [hist_old['inference_ms_old'], hist_imp['inference_ms_imp']]
    else:
        t_vals = [0.0, 0.0]   # placeholder if not yet measured
    bars   = ax.bar(p_labels, t_vals, color=colours, width=0.4)
    ax.set_title('Inference Time (ms / image)')
    ax.set_ylabel('Time (ms)')
    ax.grid(True, alpha=0.3, axis='y')
    for bar, val in zip(bars, t_vals):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.02 * (max(t_vals) + 1e-6),
                f'{val:.1f}ms', ha='center', va='bottom', fontsize=10)

    plt.tight_layout()
    plot_path = os.path.join(PLOT_DIR, 'comparison_summary.png')
    plt.savefig(plot_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"\nComparison plot saved → {plot_path}")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='Train and compare VMamba variants on Chikusei')
    parser.add_argument('--epochs',      default=5,    type=int,
                        help='Number of training epochs per model (default 5, OOM-safe)')
    parser.add_argument('--batch_size',  default=1,    type=int,
                        help='Mini-batch size (default 1, OOM-safe)')
    parser.add_argument('--d_model',     default=32,   type=int,
                        help='Feature dimension (default 32, OOM-safe; use 64 for GPU)')
    parser.add_argument('--scale',       default=4,    type=int,
                        help='Spatial upsampling scale (default 4)')
    parser.add_argument('--patch_size',  default=32,   type=int,
                        help='HR patch size passed to dataloader (default 32 → LR=8, OOM-safe; use 64 for GPU)')
    parser.add_argument('--skip_old',    action='store_true',
                        help='Skip training Old VMamba (use existing checkpoint)')
    parser.add_argument('--skip_improved', action='store_true',
                        help='Skip training Improved VMamba (use existing checkpoint)')
    args = parser.parse_args()

    # Use lighter backbone when d_model is small (OOM-safe)
    args.num_blocks = [1, 1, 1, 1] if args.d_model <= 32 else [2, 3, 3, 2]

    # Minimum HR patch for backbone: needs HR*0.25 >= 4 → HR >= 16
    min_hr = 16
    if args.patch_size < min_hr:
        print(f"[WARN] patch_size={args.patch_size} too small for backbone; raising to {min_hr}")
        args.patch_size = min_hr

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"\nDevice      : {device}")
    print(f"Dataset     : {CHIKUSEI_PATH}")
    print(f"Epochs      : {args.epochs}")
    print(f"Batch size  : {args.batch_size}")
    print(f"d_model     : {args.d_model}")
    print(f"HR patch    : {args.patch_size}×{args.patch_size}  →  LR={args.patch_size//args.scale}×{args.patch_size//args.scale}")

    # ---- Dataloaders ----
    print("\nPreparing dataloaders …")
    use_pin = device.type == 'cuda'
    train_loader, val_loader = create_dataloaders(
        mat_path    = CHIKUSEI_PATH,
        batch_size  = args.batch_size,
        patch_size  = args.patch_size,   # HR patch size
        scale       = args.scale,
        num_workers = 0,
    )
    print(f"  Train batches : {len(train_loader)}")
    print(f"  Val   batches : {len(val_loader)}")

    # ---- Excel logger ----
    timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(RESULT_DIR, f'training_log_{timestamp}.xlsx')
    logger     = ExcelLogger(excel_path) if HAS_OPENPYXL else None
    if not HAS_OPENPYXL:
        print("[WARN] Results will NOT be saved to Excel (pip install openpyxl)")

    # ---- Train Old ----
    hist_old = train_model('old', args, train_loader, val_loader, device, logger)

    # ---- Train Improved ----
    hist_imp = train_model('improved', args, train_loader, val_loader, device, logger)

    # ---- Inference time benchmark ----
    print("\nBenchmarking inference time …")
    m_old = create_vmamba_model('old',      128, 128, scale=args.scale, d_model=args.d_model, num_blocks=args.num_blocks).to(device)
    m_imp = create_vmamba_model('improved', 128, 128, scale=args.scale, d_model=args.d_model, num_blocks=args.num_blocks).to(device)

    # Load best weights
    for m, variant in [(m_old, 'old'), (m_imp, 'improved')]:
        ckpt = os.path.join(CKPT_DIR, f'{variant}_best.pth')
        if os.path.isfile(ckpt):
            st = torch.load(ckpt, map_location=device)
            if isinstance(st, dict) and 'model_state_dict' in st:
                st = st['model_state_dict']
            m.load_state_dict(st)

    t_old = measure_inference_time(m_old, device, in_ch=128, scale=args.scale)
    t_imp = measure_inference_time(m_imp, device, in_ch=128, scale=args.scale)
    hist_old['inference_ms_old'] = t_old
    hist_imp['inference_ms_imp'] = t_imp
    print(f"  Old      : {t_old:.1f} ms / image")
    print(f"  Improved : {t_imp:.1f} ms / image")

    # ---- Write Excel summary sheet ----
    if logger is not None:
        logger.write_summary(hist_old, hist_imp, t_old, t_imp)

    # ---- Plots ----
    generate_plots(hist_old, hist_imp, args)

    # ---- Summary JSON ----
    summary = {
        'config': {
            'epochs':     args.epochs,
            'batch_size': args.batch_size,
            'd_model':    args.d_model,
            'scale':      args.scale,
            'patch_size': args.patch_size,
            'dataset':    CHIKUSEI_PATH,
        },
        'old_vmamba': {
            'best_psnr':     hist_old['best_psnr'],
            'best_sam':      hist_old['best_sam'],
            'best_ergas':    hist_old['best_ergas'],
            'n_params':      hist_old['n_params'],
            'inference_ms':  t_old,
            'ckpt_path':     hist_old['ckpt_path'],
            'train_losses':  hist_old['train_losses'],
            'val_psnrs':     hist_old['val_psnrs'],
        },
        'improved_vmamba': {
            'best_psnr':     hist_imp['best_psnr'],
            'best_sam':      hist_imp['best_sam'],
            'best_ergas':    hist_imp['best_ergas'],
            'n_params':      hist_imp['n_params'],
            'inference_ms':  t_imp,
            'ckpt_path':     hist_imp['ckpt_path'],
            'train_losses':  hist_imp['train_losses'],
            'val_psnrs':     hist_imp['val_psnrs'],
        },
    }

    save_path = os.path.join(RESULT_DIR, 'summary.json')
    with open(save_path, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"Summary JSON saved → {save_path}")

    # ---- Final table ----
    print("\n" + "=" * 55)
    print("  FINAL COMPARISON")
    print("=" * 55)
    print(f"  {'Metric':<20} {'Old':>10} {'Improved':>12}")
    print("-" * 55)
    for metric, key in [('PSNR (dB)', 'best_psnr'), ('SAM (°)', 'best_sam'), ('ERGAS', 'best_ergas')]:
        o_v = hist_old[key]
        i_v = hist_imp[key]
        print(f"  {metric:<20} {o_v:>10.3f} {i_v:>12.3f}")
    print(f"  {'Params (M)':<20} {hist_old['n_params']/1e6:>10.2f} {hist_imp['n_params']/1e6:>12.2f}")
    print(f"  {'Inference (ms)':<20} {t_old:>10.1f} {t_imp:>12.1f}")
    print("=" * 55 + "\n")


if __name__ == '__main__':
    main()
